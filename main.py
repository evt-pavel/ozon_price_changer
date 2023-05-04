from openpyxl import load_workbook


template = load_workbook(filename='template.xlsx')
template_sheets_names = template.sheetnames
template_sheet = template[template_sheets_names[1]]

price = load_workbook(filename='price.xlsx', data_only = True)
price_sheets_names = price.sheetnames
price_sheet = price[price_sheets_names[1]]


# for i in range(2, template_sheet.max_row):
#     for j in range(4, price_sheet.max_row):
#         if template_sheet.cell(row=i, column=2).value == price_sheet.cell(row=j, column=3).value\
#             and price_sheet.cell(row=j, column=8).value == 0:
#             template_sheet.delete_rows(i)
#             print(template_sheet.cell(row=i, column=1).value)


for i in range(4, price_sheet.max_row):
    for j in range(2, template_sheet.max_row):
        if price_sheet.cell(row=i, column=3).value == template_sheet.cell(row=j, column=2).value\
            and price_sheet.cell(row=i, column=8).value == 0:
            template_sheet.delete_rows(j)
            print(template_sheet.cell(row=j, column=1).value)




# for i in range(2, template_sheet.max_row):
#     count = 0
#     for j in range(4, price_sheet.max_row):
#         if template_sheet.cell(row=i, column=2).value == price_sheet.cell(row=j, column=3).value:
#             template_sheet.cell(row=i, column=40).value = price_sheet.cell(row=j, column=8).value
#             template_sheet.cell(row=i, column=39).value = template_sheet.cell(row=i, column=19).value
#             count += 1
#             break

#     if count == 0:
#         print(template_sheet.cell(row=i, column=1).value)
#         template_sheet.delete_rows(i)

count = 0

for i in range(4, price_sheet.max_row):
    # print(type(price_sheet.cell(row=i, column=3).value))
    for j in range(2, template_sheet.max_row):
        # print(type(template_sheet.cell(row=j, column=1).value), template_sheet.cell(row=j, column=1).value)

        if  str(template_sheet.cell(row=j, column=2).value) == str(price_sheet.cell(row=i, column=3).value):
            print(template_sheet.cell(row=j, column=1).value)

            template_sheet.cell(row=j, column=40).value = price_sheet.cell(row=i, column=8).value
            template_sheet.cell(row=j, column=39).value = template_sheet.cell(row=j, column=19).value
            break






template.save(f'new_price.xlsx')
template.close()